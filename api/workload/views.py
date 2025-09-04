from django.shortcuts import render
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db.models import Q
from django.http import HttpResponse
from .models import Workload
from .serializers import WorkloadSerializer, WorkloadReviewSerializer
import logging
import traceback
import os
import django.db.utils
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# 获取logger
logger = logging.getLogger(__name__)

# Create your views here.

class WorkloadViewSet(viewsets.ModelViewSet):
    """工作量视图集"""
    serializer_class = WorkloadSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """获取用户可以访问的工作量列表"""
        user = self.request.user
        submitted = self.request.query_params.get('submitted', 'false').lower() == 'true'

        # 如果是获取已提交的工作量列表，只返回自己提交的
        if submitted:
            return Workload.objects.filter(submitter=user)

        # 否则根据角色返回可访问的工作量
        if user.role == 'student':
            # 学生只能看到自己提交的工作量
            return Workload.objects.filter(submitter=user)
        elif user.role == 'mentor':
            # 导师可以看到：1.自己提交的工作量 2.自己需要审核的学生工作量
            return Workload.objects.filter(
                Q(submitter=user) |  # 自己提交的
                Q(mentor_reviewer=user, submitter__role='student')  # 需要审核的学生工作量
            )
        elif user.role == 'teacher':
            # 教师可以看到所有工作量
            return Workload.objects.all()
        return Workload.objects.none()

    def create(self, request, *args, **kwargs):
        """创建工作量"""
        try:
            # 记录请求信息
            logger.info(f"用户 {request.user.username} 正在提交工作量")
            
            # 检查文件上传
            if 'attachments' in request.FILES:
                file = request.FILES['attachments']
                logger.info(f"文件上传: {file.name}, 大小: {file.size} 字节, 类型: {file.content_type}")
                
                # 检查上传目录是否存在
                upload_dir = os.path.join('media', 'workload_files')
                if not os.path.exists(upload_dir):
                    os.makedirs(upload_dir, exist_ok=True)
                    logger.info(f"创建上传目录: {upload_dir}")

            # 调用序列化器进行验证
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # 保存工作量
            try:
                self.perform_create(serializer)
            except django.db.utils.IntegrityError as e:
                # 捕获主键冲突错误
                if "Duplicate entry" in str(e) and "PRIMARY" in str(e):
                    # 从错误消息中提取ID
                    match = re.search(r"Duplicate entry '(\d+)' for key", str(e))
                    if match:
                        conflict_id = match.group(1)
                        logger.warning(f"主键冲突: ID {conflict_id}，尝试使用不同的ID创建")
                        
                        # 获取最大ID
                        max_id = Workload.objects.all().order_by('-id').first()
                        new_id = (max_id.id + 1) if max_id else 1
                        
                        # 尝试用新ID创建
                        serializer.save(id=new_id, submitter=request.user)
                else:
                    # 其他完整性错误，重新抛出
                    raise
            
            # 记录成功信息
            logger.info(f"工作量提交成功: ID {serializer.instance.id}, 名称: {serializer.instance.name}")
            
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            # 记录错误信息
            logger.error(f"工作量提交失败: {str(e)}")
            logger.error(traceback.format_exc())
            
            # 返回友好的错误信息
            if hasattr(e, 'detail'):
                error_detail = str(e.detail)
            else:
                error_detail = str(e)
                
            return Response(
                {"detail": f"提交失败: {error_detail}"},
                status=status.HTTP_400_BAD_REQUEST
            )

    def perform_create(self, serializer):
        """创建工作量时设置提交者"""
        try:
            serializer.save(submitter=self.request.user)
        except Exception as e:
            logger.error(f"保存工作量失败: {str(e)}")
            logger.error(traceback.format_exc())
            raise e

    def update(self, request, *args, **kwargs):
        """更新工作量时进行权限和状态检查"""
        try:
            instance = self.get_object()
            user = self.request.user
            
            # 重新从数据库获取最新状态
            instance.refresh_from_db()
            
            # 检查工作量当前状态
            if user.role != 'teacher':  # 只对非教师角色进行状态检查
                if instance.status not in ['pending', 'mentor_rejected', 'teacher_rejected']:
                    return Response(
                        {"detail": "只能修改未审核或审核未通过的工作量"},
                        status=status.HTTP_403_FORBIDDEN
                    )

            # 检查文件上传
            if 'attachments' in request.FILES:
                file = request.FILES['attachments']
                logger.info(f"文件更新: {file.name}, 大小: {file.size} 字节, 类型: {file.content_type}")

            # 调用父类的 update 方法
            serializer = self.get_serializer(instance, data=request.data, partial=kwargs.get('partial', False))
            serializer.is_valid(raise_exception=True)
            
            # 如果是被驳回的工作量，修改后重置状态为待审核
            if user.role != 'teacher' and instance.status in ['mentor_rejected', 'teacher_rejected']:
                serializer.save(status='pending')
            else:
                serializer.save()

            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"更新工作量失败: {str(e)}")
            logger.error(traceback.format_exc())
            
            if hasattr(e, 'detail'):
                error_detail = str(e.detail)
            else:
                error_detail = str(e)
                
            return Response(
                {"detail": f"更新失败: {error_detail}"},
                status=status.HTTP_400_BAD_REQUEST
            )

    def perform_update(self, serializer):
        """更新工作量的实际操作"""
        serializer.save()

    def perform_destroy(self, instance):
        """删除工作量时进行权限检查"""
        if instance.status not in ['pending', 'mentor_rejected', 'teacher_rejected']:
            return Response(
                {"detail": "只能删除未审核或审核未通过的工作量"},
                status=status.HTTP_403_FORBIDDEN
            )
        instance.delete()

    @action(detail=False, methods=['get'])
    def pending_review(self, request):
        """获取待审核的工作量列表"""
        user = request.user
        if user.role == 'mentor':
            # 导师获取自己需要审核的待审核工作量
            queryset = Workload.objects.filter(
                mentor_reviewer=user,
                submitter__role='student'
            ).filter(
                Q(status='pending') # 待审核的工作量
            )
        elif user.role == 'teacher':
            # 教师获取所有需要教师审核的工作量
            queryset = Workload.objects.filter(
                Q(status='mentor_approved') |  # 导师已审核通过的工作量
                Q(submitter__role='mentor', status='pending')  # 导师提交的待审核工作量
            )
        else:
            return Response(
                {"detail": "只有导师和教师可以查看待审核工作量"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def reviewed(self, request):
        """获取已审核的工作量列表"""
        user = request.user
        if user.role == 'mentor':
            # 导师获取自己已审核的工作量
            queryset = Workload.objects.filter(
                mentor_reviewer=user,
                submitter__role='student'
            ).exclude(status='pending')
        elif user.role == 'teacher':
            # 教师获取所有已审核的工作量
            queryset = Workload.objects.filter(
                teacher_reviewer=user)
        else:
            return Response(
                {"detail": "只有导师和教师可以查看已审核工作量列表"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def all_workloads(self, request):
        """教师获取所有工作量列表"""
        user = request.user
        if user.role != 'teacher':
            return Response(
                {"detail": "只有教师可以查看所有工作量"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        queryset = Workload.objects.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        """审核工作量"""
        instance = self.get_object()
        user = request.user

        # 检查审核权限
        if user.role == 'mentor':
            if instance.submitter.role != 'student' or instance.mentor_reviewer != user:
                return Response(
                    {"detail": "您不是该工作量的指定审核导师"},
                    status=status.HTTP_403_FORBIDDEN
                )
        elif user.role == 'teacher':
            if instance.submitter.role == 'student' and instance.status != 'mentor_approved':
                return Response(
                    {"detail": "学生提交的工作量需要导师审核通过后才能进行教师审核"},
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            return Response(
                {"detail": "只有导师和教师可以审核工作量"},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = WorkloadReviewSerializer(
            instance,
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            # 返回更新后的完整工作量信息
            return Response(
                self.get_serializer(instance).data
            )
        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=False, methods=['post'])
    def export(self, request):
        """导出选中的工作量为Excel文件"""
        user = request.user
        if user.role != 'teacher':
            return Response(
                {"detail": "只有教师可以导出工作量"},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            # 获取要导出的工作量ID列表
            workload_ids = request.data.get('workload_ids', [])
            if not workload_ids:
                return Response(
                    {"detail": "请选择要导出的工作量"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 获取工作量数据
            workloads = Workload.objects.filter(id__in=workload_ids)
            if not workloads:
                return Response(
                    {"detail": "未找到指定的工作量"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "工作量记录"

            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 设置表头
            headers = [
                '提交人', '工作量名称', '工作量内容', '工作来源', '工作类型',
                '开始日期', '结束日期', '工作强度类型', '工作强度值', '状态',
                '导师评语', '导师审核时间', '教师评语', '教师审核时间',
                '创建时间', '大创阶段', '参与人及占比', '已发助教工资'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 写入数据
            for row, workload in enumerate(workloads, 2):
                ws.cell(row=row, column=1, value=workload.submitter.username)
                ws.cell(row=row, column=2, value=workload.name)
                ws.cell(row=row, column=3, value=workload.content)
                ws.cell(row=row, column=4, value=workload.get_source_display())
                ws.cell(row=row, column=5, value=workload.get_work_type_display())
                ws.cell(row=row, column=6, value=workload.start_date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=7, value=workload.end_date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=8, value=workload.get_intensity_type_display())
                ws.cell(row=row, column=9, value=workload.intensity_value)
                ws.cell(row=row, column=10, value=workload.get_status_display())
                ws.cell(row=row, column=11, value=workload.mentor_comment)
                ws.cell(row=row, column=12, value=workload.mentor_review_time.strftime('%Y-%m-%d %H:%M:%S') if workload.mentor_review_time else '')
                ws.cell(row=row, column=13, value=workload.teacher_comment)
                ws.cell(row=row, column=14, value=workload.teacher_review_time.strftime('%Y-%m-%d %H:%M:%S') if workload.teacher_review_time else '')
                ws.cell(row=row, column=15, value=workload.created_at.strftime('%Y-%m-%d %H:%M:%S'))

                # 大创阶段（仅当来源为大创）
                ws.cell(row=row, column=16,
                        value=workload.get_source_display() == '大创' and workload.get_innovation_stage_display() or '')

                # 大创参与人及占比
                if workload.source == 'innovation':
                    share_text = "; ".join(
                        [f"{share.user.username}:{share.percentage}%" for share in workload.shares.all()]
                    )
                    ws.cell(row=row, column=17, value=share_text)
                else:
                    ws.cell(row=row, column=17, value='')

                # 已发助教工资（仅当来源为助教）
                ws.cell(row=row, column=18,
                        value=workload.assistant_salary_paid if workload.source == 'assistant' else '')

            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # 创建响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=workload_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            # 保存文件
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"导出工作量失败: {str(e)}")
            logger.error(traceback.format_exc())
            return Response(
                {"detail": f"导出失败: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
